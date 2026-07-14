#!/usr/bin/env python3
"""Import GBM Main (MXN) + GBM USA (USD) Excel portfolio exports into broker_positions
and rebuild consolidated positions table."""
from __future__ import annotations

import os
import re
import sys
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

sys.path.insert(0, str(Path.home() / ".hermes" / "scripts"))
try:
    import hermes_secrets_bootstrap  # noqa: F401
except Exception:
    pass

HERMES = Path.home() / ".hermes"
MAIN_XLSX = HERMES / "cache/documents/doc_ce4382e6f94b_App_GBM_Detalle_Portafolio__1783994709273.xlsx"
USA_XLSX = HERMES / "cache/documents/doc_c47306f0ab97_App_GBM_Detalle_Portafolio_USA_1783994719028.xlsx"

SKIP_TICKERS = {
    "EMISORA/FONDO",
    "EFEC. MISMO DIA",
    "EFEC.  MISMO DIA",
    "EFEC. 24 HRS.",
    "EFEC. 48 HRS.",
    "EFEC. MAYOR 48 HRS.",
    "EFECTIVO",
    "EFECTIVO",
    "LIQUIDEZ",
    "EFECTIVO",
}


def load_env() -> None:
    envp = HERMES / ".env"
    if not envp.exists():
        return
    for line in envp.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def connect():
    import psycopg2

    load_env()
    pw = os.environ.get("DB_PASSWORD") or os.environ.get("PGPASSWORD") or ""
    if len(pw) < 5:
        raise RuntimeError("DB password missing")
    return psycopg2.connect(
        host=os.environ.get("PGHOST", "acela.proxy.rlwy.net"),
        port=int(os.environ.get("PGPORT", "35577")),
        dbname=os.environ.get("PGDATABASE", "railway"),
        user=os.environ.get("PGUSER", "postgres"),
        password=pw,
        connect_timeout=20,
    )


def money(val: Any) -> Optional[float]:
    if val is None or val == "-" or val == "":
        return None
    if isinstance(val, (int, float, Decimal)):
        return float(val)
    s = str(val).strip()
    if s in ("-", "", "None"):
        return None
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def normalize_ticker(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    t = str(raw).strip()
    if not t or t.upper() in SKIP_TICKERS:
        return None
    # Section headers
    if t.startswith("Mercado") or t.startswith("Valores") or t.startswith("Efectivo") or t.startswith("Liquidez"):
        return None
    t = t.replace(" *", "").replace(" N", "").replace("*", "").strip()
    t = re.sub(r"\s+", " ", t)
    if t.upper() in ("EFECTIVO", "EFECTIVO USD", "CASH"):
        return None
    if t.lower() == "efectivo":
        return None
    # Reporto / money market keep as-is but tag later
    return t


def parse_gbm_sheet(path: Path, broker: str, currency: str, mxn_rate: float) -> Tuple[List[Dict], Dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    positions: List[Dict] = []
    cash_mxn = 0.0
    cash_usd = 0.0
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        raw = str(row[0]).strip()
        # cash rows
        if raw.upper().startswith("EFEC") or raw.lower() == "efectivo":
            val = money(row[5])
            if val:
                if currency == "MXN":
                    cash_mxn += val
                    cash_usd += val / mxn_rate
                else:
                    cash_usd += val
            continue
        ticker = normalize_ticker(raw)
        if not ticker:
            continue
        # header repeat
        if ticker.upper() == "EMISORA/FONDO":
            continue
        shares = money(row[1])
        avg_cost = money(row[2])
        live_price = money(row[3])
        live_value = money(row[5])
        if shares is None and live_value is None:
            continue
        # skip zero value dust without shares
        if (shares or 0) == 0 and (live_value or 0) == 0:
            continue
        # Reporto / cash-like with no equity price skip from equity book if tiny money market?
        # Keep GBM O, BI reporto as holdings
        if currency == "MXN":
            live_value_usd = (live_value or 0) / mxn_rate
            live_price_usd = (live_price or 0) / mxn_rate if live_price else None
            avg_cost_usd = (avg_cost or 0) / mxn_rate if avg_cost else None
            live_value_native = live_value
            live_price_native = live_price
            avg_cost_native = avg_cost
        else:
            live_value_usd = live_value or 0
            live_price_usd = live_price
            avg_cost_usd = avg_cost
            live_value_native = live_value
            live_price_native = live_price
            avg_cost_native = avg_cost

        positions.append(
            {
                "broker": broker,
                "ticker": ticker,
                "shares": shares or 0,
                "avg_cost": avg_cost_native,
                "live_price": live_price_native,
                "live_value": live_value_native,
                "live_value_usd": live_value_usd,
                "currency": currency,
                "avg_cost_usd": avg_cost_usd,
                "live_price_usd": live_price_usd,
            }
        )
    meta = {"cash_usd": cash_usd, "cash_mxn": cash_mxn, "count": len(positions)}
    return positions, meta


def get_mxn_rate() -> float:
    try:
        import yfinance as yf

        h = yf.Ticker("USDMXN=X").history(period="5d")
        if not h.empty:
            return float(h["Close"].iloc[-1])
    except Exception:
        pass
    return 17.50


def upsert_broker(conn, positions: List[Dict], broker: str) -> None:
    cur = conn.cursor()
    # Replace broker snapshot: delete all then insert (clean)
    cur.execute("DELETE FROM broker_positions WHERE broker = %s", (broker,))
    now = datetime.now(timezone.utc)
    for p in positions:
        # Prefer USD price for live_price when currency MXN for cross-broker consistency?
        # Keep native price in live_price; store USD value in live_value_usd
        cur.execute(
            """
            INSERT INTO broker_positions (
                broker, ticker, shares, avg_cost, live_price, live_value,
                currency, live_value_usd, source, last_sync_at, created_at, updated_at, price_source
            ) VALUES (
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
            """,
            (
                broker,
                p["ticker"],
                p["shares"],
                p["avg_cost"],
                p["live_price"],
                p["live_value"],
                p["currency"],
                p["live_value_usd"],
                "gbm_excel_import",
                now,
                now,
                now,
                "gbm_export",
            ),
        )
    conn.commit()
    print(f"  ✅ {broker}: {len(positions)} rows written")


def rebuild_positions(conn) -> Dict[str, Any]:
    """Aggregate all broker_positions into positions (sum shares/value, brokers array)."""
    cur = conn.cursor()
    # Grades map from latest unified / existing positions
    cur.execute(
        """
        SELECT ticker, unified_grade, action FROM unified_grades
        """
    )
    ug = {r[0]: (r[1], r[2]) for r in cur.fetchall()}
    cur.execute("SELECT ticker, grade, council, sector FROM positions")
    old = {r[0]: {"grade": r[1], "council": r[2], "sector": r[3]} for r in cur.fetchall()}

    cur.execute(
        """
        SELECT broker, ticker, shares, avg_cost, live_price, live_value, live_value_usd, currency
        FROM broker_positions
        WHERE COALESCE(shares, 0) > 0 OR COALESCE(live_value_usd, 0) > 0
        """
    )
    rows = cur.fetchall()
    by_ticker: Dict[str, Dict[str, Any]] = {}
    for broker, ticker, shares, avg_cost, live_price, live_value, live_value_usd, currency in rows:
        t = ticker
        # Normalize crypto-like display already
        rec = by_ticker.setdefault(
            t,
            {
                "shares": 0.0,
                "value_usd": 0.0,
                "brokers": set(),
                "prices": [],
                "avg_costs": [],
                "currencies": set(),
            },
        )
        rec["shares"] += float(shares or 0)
        rec["value_usd"] += float(live_value_usd or 0)
        rec["brokers"].add(broker)
        if live_price is not None:
            # Only use USD-ish prices for live_price if currency USD; for MXN convert
            if (currency or "").upper() == "MXN":
                # approximate USD price if shares>0
                if shares and float(shares) > 0 and live_value_usd:
                    rec["prices"].append(float(live_value_usd) / float(shares))
            else:
                rec["prices"].append(float(live_price))
        if avg_cost is not None and (currency or "").upper() != "MXN":
            rec["avg_costs"].append(float(avg_cost))
        rec["currencies"].add(currency or "USD")

    # Zero out / replace positions content carefully:
    # Keep MIRROR_TOTAL if present from eToro
    cur.execute("SELECT ticker, shares, live_value_usd, grade, council, brokers FROM positions WHERE ticker = 'MIRROR_TOTAL'")
    mirror = cur.fetchone()

    cur.execute("DELETE FROM positions WHERE ticker IS DISTINCT FROM 'MIRROR_TOTAL'")

    now = datetime.now(timezone.utc)
    inserted = 0
    for t, rec in sorted(by_ticker.items(), key=lambda x: -x[1]["value_usd"]):
        if t == "MIRROR_TOTAL":
            continue
        shares = rec["shares"]
        value = rec["value_usd"]
        if value <= 0 and shares <= 0:
            continue
        price = sum(rec["prices"]) / len(rec["prices"]) if rec["prices"] else (value / shares if shares else None)
        # if price still None and shares and value
        if price is None and shares:
            price = value / shares
        avg = sum(rec["avg_costs"]) / len(rec["avg_costs"]) if rec["avg_costs"] else None
        grade = None
        council = None
        sector = None
        if t in ug:
            grade = int(float(ug[t][0])) if ug[t][0] is not None else None
            council = ug[t][1]
        elif t in old:
            grade = old[t]["grade"]
            council = old[t]["council"]
            sector = old[t]["sector"]
        brokers = sorted(rec["brokers"])
        # primary currency USD for consolidated
        cur.execute(
            """
            INSERT INTO positions (
                ticker, shares, avg_cost, live_price, live_value, grade, council, brokers,
                sector, updated_at, currency, live_value_usd, price_source, mode
            ) VALUES (
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
            ON CONFLICT (ticker) DO UPDATE SET
                shares = EXCLUDED.shares,
                avg_cost = EXCLUDED.avg_cost,
                live_price = EXCLUDED.live_price,
                live_value = EXCLUDED.live_value,
                grade = COALESCE(EXCLUDED.grade, positions.grade),
                council = COALESCE(EXCLUDED.council, positions.council),
                brokers = EXCLUDED.brokers,
                updated_at = EXCLUDED.updated_at,
                currency = EXCLUDED.currency,
                live_value_usd = EXCLUDED.live_value_usd,
                price_source = EXCLUDED.price_source
            """,
            (
                t,
                shares,
                avg,
                price,
                value,  # live_value in USD terms for consolidated
                grade,
                council,
                brokers,
                sector,
                now,
                "USD",
                value,
                "broker_aggregate",
                "live",
            ),
        )
        inserted += 1

    # restore mirror if was deleted wrongly - we excluded it from delete? re-check
    if mirror:
        # ensure still exists
        cur.execute("SELECT 1 FROM positions WHERE ticker='MIRROR_TOTAL'")
        if not cur.fetchone():
            cur.execute(
                """
                INSERT INTO positions (ticker, shares, live_value_usd, live_value, grade, council, brokers, currency, updated_at, mode)
                VALUES ('MIRROR_TOTAL', %s, %s, %s, %s, %s, %s, 'USD', %s, 'mirror')
                """,
                (mirror[1], mirror[2], mirror[2], mirror[3], mirror[4], mirror[5], now),
            )

    conn.commit()
    cur.execute(
        "SELECT COUNT(*), COALESCE(SUM(live_value_usd),0) FROM positions WHERE COALESCE(shares,0)>0 OR ticker='MIRROR_TOTAL'"
    )
    c, aum = cur.fetchone()
    cur.execute(
        """
        SELECT broker, COUNT(*), COALESCE(SUM(live_value_usd),0)
        FROM broker_positions GROUP BY broker ORDER BY 3 DESC
        """
    )
    by_broker = cur.fetchall()
    return {"positions": int(c), "aum": float(aum), "by_broker": by_broker, "inserted": inserted}


def main() -> int:
    if not MAIN_XLSX.exists() or not USA_XLSX.exists():
        print("Missing Excel files")
        print("MAIN", MAIN_XLSX.exists(), MAIN_XLSX)
        print("USA", USA_XLSX.exists(), USA_XLSX)
        return 1

    mxn = get_mxn_rate()
    print(f"MXN/USD rate: {mxn:.4f}")

    main_pos, main_meta = parse_gbm_sheet(MAIN_XLSX, "GBM Main", "MXN", mxn)
    usa_pos, usa_meta = parse_gbm_sheet(USA_XLSX, "GBM USA", "USD", mxn)

    main_usd = sum(p["live_value_usd"] for p in main_pos) + main_meta["cash_usd"]
    usa_usd = sum(p["live_value_usd"] for p in usa_pos) + usa_meta["cash_usd"]

    print(f"\nGBM Main: {len(main_pos)} positions, equity ${sum(p['live_value_usd'] for p in main_pos):,.0f} + cash ${main_meta['cash_usd']:.2f} ≈ ${main_usd:,.0f}")
    for p in sorted(main_pos, key=lambda x: -x["live_value_usd"])[:10]:
        print(f"  {p['ticker']:10} sh={p['shares']:>10}  ${p['live_value_usd']:>10,.0f}")
    print(f"\nGBM USA: {len(usa_pos)} positions, equity ${sum(p['live_value_usd'] for p in usa_pos):,.0f} + cash ${usa_meta['cash_usd']:.2f} ≈ ${usa_usd:,.0f}")
    for p in sorted(usa_pos, key=lambda x: -x["live_value_usd"]):
        print(f"  {p['ticker']:10} sh={p['shares']:>12.4f}  ${p['live_value_usd']:>10,.0f}")

    conn = connect()
    print("\nWriting broker_positions...")
    upsert_broker(conn, main_pos, "GBM Main")
    upsert_broker(conn, usa_pos, "GBM USA")

    # Optional cash rows as CASH tickers? skip — cash tiny on USA, tiny on main
    print("\nRebuilding consolidated positions...")
    summary = rebuild_positions(conn)
    print(f"  positions rows: {summary['inserted']} (count query {summary['positions']})")
    print(f"  consolidated AUM (positions): ${summary['aum']:,.0f}")
    print("  by broker:")
    for b, n, v in summary["by_broker"]:
        print(f"    {b:12} {n:3}  ${float(v):>12,.0f}")

    conn.close()
    print("\n✅ GBM Main + USA import complete")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
